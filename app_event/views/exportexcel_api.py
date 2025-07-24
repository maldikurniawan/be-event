import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import permissions
from rest_framework.views import APIView

from app_event.models import EventAttendance
from be_event.permissions import PermissionMixin


class EventAttendanceExportExcelApi(PermissionMixin, APIView):
    permission_classes = (permissions.AllowAny,)

    def get(self, request, slug):
        # Ambil data attendance berdasarkan slug
        attendance = EventAttendance.objects.filter(event__slug=slug).order_by(
            "-created_at"
        )

        # Buat workbook & sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Event Attendance"

        # Header kolom
        headers = [
            "Nama",
            "No HP",
            "Email",
            "Perusahaan",
            "IP Address",
            "Tanggal",
        ]
        ws.append(headers)

        # Set bold untuk header
        for col_idx, _ in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

        # Isi data
        for row_idx, item in enumerate(attendance, start=2):
            ws.cell(row=row_idx, column=1, value=item.nama)
            ws.cell(row=row_idx, column=2, value=item.nohp)
            ws.cell(row=row_idx, column=3, value=item.email)
            ws.cell(row=row_idx, column=4, value=item.nama_perusahaan)
            ws.cell(row=row_idx, column=5, value=item.ip_address)
            ws.cell(
                row=row_idx,
                column=6,
                value=item.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            )

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        # Buat response Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"attendance-{slug}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)

        return response
